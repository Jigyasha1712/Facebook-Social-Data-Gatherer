from facebook_scraper import get_posts
import openpyxl
import time
import random
from datetime import datetime

excel_file = r"test.xlsx"

# Load Excel
wb = openpyxl.load_workbook(excel_file)
ws = wb.active

url_column_name = "Post URL"

# Find URL column
url_col_idx = None
for col in range(1, ws.max_column + 1):
    if ws.cell(row=1, column=col).value == url_column_name:
        url_col_idx = col

if not url_col_idx:
    raise Exception("Post URL column not found!")

# Result headers
headers = ["Likes", "Comments", "Shares", "Text", "Date", "Time", "Username"]

start_col = ws.max_column + 1
for i, h in enumerate(headers):
    ws.cell(row=1, column=start_col + i, value=h)

cols = {h: start_col+i for i, h in enumerate(headers)}

total_rows = ws.max_row - 1
processed = 0

# Main loop
for row in range(2, ws.max_row + 1):
    url = ws.cell(row=row, column=url_col_idx).value
    if not url:
        continue

    try:
        post_data = None
        for post in get_posts(
            post_urls=[url],
            cookies='Cookies.txt'   # <<< very important
        ):
            post_data = post
            break
        
        if not post_data:
            print(f"No data: {url}")
            continue

        dt = post_data.get("time")

        ws.cell(row=row, column=cols["Likes"], value=post_data.get("likes"))
        ws.cell(row=row, column=cols["Comments"], value=post_data.get("comments"))
        ws.cell(row=row, column=cols["Shares"], value=post_data.get("shares"))
        ws.cell(row=row, column=cols["Text"], value=post_data.get("text"))
        ws.cell(row=row, column=cols["Date"], value=dt.strftime("%d-%b-%Y") if dt else "")
        ws.cell(row=row, column=cols["Time"], value=dt.strftime("%H:%M:%S") if dt else "")
        ws.cell(row=row, column=cols["Username"], value=post_data.get("username"))

        processed += 1

        print(f"✅ Row {row} done: {post_data.get('username')}")

    except Exception as e:
        print(f"❌ Error row {row}: {e}")

    time.sleep(random.choice([3,4,5,6]))

wb.save(excel_file)
print(f"\nDone: {processed}/{total_rows}")
